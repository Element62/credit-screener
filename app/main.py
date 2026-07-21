from __future__ import annotations

from io import BytesIO
import shutil
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from fastapi import Body, Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

from .auth import SESSION_COOKIE, create_session_token, decode_session_token
from .config import DATA_DIR, load_settings
from .data_loader import WorkbookData, load_workbook
from .db import get_all_coverage, init_db, save_coverage
from .market_data import load_market_data
from .report_processor import ReportProcessorService, normalize_company_name


app = FastAPI(title="Credit Screenr")
settings = load_settings()


@app.on_event("startup")
def startup() -> None:
    init_db()
DATA_DIR.mkdir(exist_ok=True)
settings.processed_report_dir.mkdir(parents=True, exist_ok=True)


class DataStore:
    def __init__(self) -> None:
        self.dataset: WorkbookData | None = None

    def refresh(self, workbook_path: Path) -> WorkbookData:
        self.dataset = load_workbook(workbook_path)
        return self.dataset


store = DataStore()
report_service = ReportProcessorService(
    api_key=settings.openai_api_key,
    report_dir=settings.report_dir,
    processed_dir=settings.processed_report_dir,
    summary_workbook_path=settings.report_summary_workbook,
    ca_bundle=settings.openai_ca_bundle or None,
)


def get_current_user(request: Request) -> str:
    token = request.cookies.get(SESSION_COOKIE)
    session = decode_session_token(settings.app_secret, token)
    if not session:
        raise HTTPException(status_code=401, detail="Unauthorized")
    return session.username


def ensure_data_loaded() -> WorkbookData:
    if store.dataset is None:
        workbook_path = settings.workbook_path
        if not workbook_path.exists():
            raise HTTPException(status_code=500, detail=f"Workbook not found: {workbook_path}")
        store.refresh(workbook_path)
    return store.dataset


def _build_coverage_map(dataset: WorkbookData) -> dict:
    """Industry defaults from workbook, overridden by any DB-saved entries."""
    coverage: dict = {}
    for row in dataset.issuer_table:
        ticker = row.get("PARENT_TICKER")
        industry = row.get("Industry")
        if ticker and industry and industry in dataset.analyst_map:
            coverage[ticker] = {
                "primary": list(dataset.analyst_map[industry]["primary"]),
                "secondary": list(dataset.analyst_map[industry]["secondary"]),
            }
    coverage.update(get_all_coverage())  # DB entries override industry defaults
    return coverage


def enrich_issuers_with_reports(dataset: WorkbookData) -> dict:
    report_records = report_service.snapshot()
    issuers = []
    for row in dataset.issuer_table:
        enriched = dict(row)
        report = report_records.get(normalize_company_name(str(row.get("Issuer", ""))))
        if report:
            enriched["REPORT_SENTIMENT_SCORE"] = report.get("sentiment_score")
            enriched["REPORT_SENTIMENT_LABEL"] = report.get("sentiment_label")
            enriched["REPORT_SENTIMENT_COLOR"] = report.get("sentiment_color")
            enriched["REPORT_SUMMARY_BULLETS"] = report.get("summary_bullets", [])
            enriched["REPORT_DATE"] = report.get("report_date")
            enriched["REPORT_SOURCE_FILE"] = report.get("source_file")
        else:
            enriched["REPORT_SENTIMENT_SCORE"] = None
            enriched["REPORT_SENTIMENT_LABEL"] = ""
            enriched["REPORT_SENTIMENT_COLOR"] = ""
            enriched["REPORT_SUMMARY_BULLETS"] = []
            enriched["REPORT_DATE"] = None
            enriched["REPORT_SOURCE_FILE"] = None
        issuers.append(enriched)
    return {
        "issuers": issuers,
        "filters": {**dataset.filters, "analyst_names": dataset.analyst_names},
        "metadata": {**dataset.metadata, "processed_report_count": len(report_records)},
    }


def build_excel_response(rows: list[dict], filename: str, sheet_name: str) -> StreamingResponse:
    output = BytesIO()
    df = pd.DataFrame(rows)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name[:31] or "Sheet1")
    output.seek(0)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


# ---------------------------------------------------------------------------
# Formatted issuer export
# ---------------------------------------------------------------------------
_NAVY = "113B74"
_WHITE = "FFFAF4"
_BG = "F6F7F4"
_GREEN = "0B7F5A"
_RED = "9F3E32"
_AMBER = "B58A17"
_LINE = "D0D5DD"
_GROUP_BORDER_COLOR = "6B8CBF"

_THIN = Side(style="thin", color=_LINE)
_GROUP = Side(style="medium", color=_GROUP_BORDER_COLOR)
_NO = Side(style=None)

_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_ALT_FILL = PatternFill("solid", fgColor="EFF2F8")
_HOLDING_FILL = PatternFill("solid", fgColor="DFF2E9")
_HOLDING_ALT_FILL = PatternFill("solid", fgColor="C8E8D5")

_ISSUER_COLS = [
    "Issuer", "Sector",
    "Secured Face ($BN)", "Unsecured Face ($BN)", "Preferred Face ($BN)",
    "52W PEAK UPSIDE SECURED ($MM)", "52W PEAK UPSIDE UNSECURED ($MM)", "52W PEAK UPSIDE PREFERRED ($MM)",
    "Price", "Yield", "3M Price Move", "7D Price Move",
]

_COL_LABELS = [
    "Issuer", "Sector",
    "Secured", "Unsecured", "Preferred",
    "Secured", "Unsecured", "Preferred",
    "Price", "Yield", "3M Move", "7D Move",
]

# (start_col_1based, end_col_1based, label)
_GROUPS = [
    (1, 1, "Issuer"),
    (2, 2, "Sector"),
    (3, 5, "FACE (In $ Billions)"),
    (6, 8, "52W PEAK UPSIDE (In $ Millions)"),
    (9, 9, "Price"),
    (10, 10, "Yield"),
    (11, 12, "Price Move"),
]

# Columns that get left group border (1-based)
_GROUP_LEFT = {3, 6}
# Columns that get right group border (1-based)
_GROUP_RIGHT = {1, 5, 8}


def _group_border(col: int, base_border: Border) -> Border:
    left = _GROUP if col in _GROUP_LEFT else base_border.left
    right = _GROUP if col in _GROUP_RIGHT else base_border.right
    return Border(left=left, right=right, top=base_border.top, bottom=base_border.bottom)


def build_issuer_excel_response(rows: list[dict], upside_mode: str) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Issuer Screen"

    upside_label = "52W PEAK UPSIDE (In $ Millions)" if upside_mode == "52w" else "RETURN TO PAR UPSIDE (In $ Millions)"
    upside_keys = (
        ["52W PEAK UPSIDE SECURED ($MM)", "52W PEAK UPSIDE UNSECURED ($MM)", "52W PEAK UPSIDE PREFERRED ($MM)"]
        if upside_mode == "52w"
        else ["RETURN TO PAR SECURED ($MM)", "RETURN TO PAR UNSECURED ($MM)", "RETURN TO PAR PREFERRED ($MM)"]
    )
    col_keys = [
        "Issuer", "Sector",
        "Secured Face ($BN)", "Unsecured Face ($BN)", "Preferred Face ($BN)",
        *upside_keys,
        "Price", "Yield", "3M Price Move", "7D Price Move", "3M MV Change ($MM)",
    ]
    n_cols = len(col_keys)

    header_font = Font(name="Calibri", bold=True, color=_WHITE, size=9)
    header_font_italic = Font(name="Calibri", bold=True, italic=True, color=_WHITE, size=9)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")

    # ── Row 1: group headers ──────────────────────────────────────────────
    groups = [
        (1, 1, "Issuer", False),
        (2, 2, "Sector", False),
        (3, 5, "FACE", True),          # italic units appended below via rich-text workaround
        (6, 8, upside_label, True),
        (9, 9, "Price", False),
        (10, 10, "Yield", False),
        (11, 12, "Price Move", False),
        (13, 13, "3M MV Change (In $MM)", False),
    ]
    for start, end, label, _italic in groups:
        cell = ws.cell(row=1, column=start, value=label)
        cell.font = header_font
        cell.fill = _HEADER_FILL
        cell.alignment = center
        if end > start:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        b = _group_border(start, Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN))
        cell.border = b

    # fill remaining merged cells in row 1
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=c)
        if cell.value is None:
            cell.fill = _HEADER_FILL
            cell.border = _group_border(c, Border(top=_THIN, bottom=_THIN, left=_NO, right=_NO))

    # ── Row 2: column labels ──────────────────────────────────────────────
    col_labels = [
        "Issuer", "Sector",
        "Secured", "Unsecured", "Preferred",
        "Secured", "Unsecured", "Preferred",
        "Price", "Yield", "3M Move", "7D Move", "3M MV Chg",
    ]
    for c, label in enumerate(col_labels, start=1):
        cell = ws.cell(row=2, column=c, value=label)
        cell.font = header_font
        cell.fill = _HEADER_FILL
        cell.alignment = center
        cell.border = _group_border(c, Border(top=_THIN, bottom=Side(style="medium", color=_NAVY), left=_THIN, right=_THIN))

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 18

    # ── Data rows ────────────────────────────────────────────────────────
    data_font = Font(name="Calibri", size=9)
    data_font_bold = Font(name="Calibri", size=9, bold=True)

    for r_idx, row in enumerate(rows, start=3):
        is_holding = bool(row.get("_HAS_HOLDING"))
        if is_holding:
            fill = _HOLDING_ALT_FILL if r_idx % 2 == 0 else _HOLDING_FILL
        else:
            fill = _ALT_FILL if r_idx % 2 == 0 else PatternFill("solid", fgColor=_WHITE)
        for c_idx, key in enumerate(col_keys, start=1):
            value = row.get(key)
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.fill = fill
            cell.border = _group_border(c_idx, Border(bottom=_THIN, left=_NO, right=_NO, top=_NO))

            if key in ("3M Price Move", "7D Price Move") and value is not None:
                try:
                    v = float(value)
                    cell.value = round(v, 2)
                    cell.font = Font(name="Calibri", size=9, bold=True,
                                     color=_GREEN if v > 0 else (_RED if v < 0 else "5D6A7D"))
                    cell.number_format = '+0.00;-0.00;0.00'
                    cell.alignment = right_align
                    continue
                except (TypeError, ValueError):
                    pass

            if key in ("Price", "Yield") and value is not None:
                try:
                    cell.value = round(float(value), 2)
                    cell.number_format = '0.00'
                    cell.font = data_font
                    cell.alignment = right_align
                    continue
                except (TypeError, ValueError):
                    pass

            if key in ("Secured Face ($BN)", "Unsecured Face ($BN)", "Preferred Face ($BN)") and value is not None:
                try:
                    cell.value = round(float(value), 2)
                    cell.number_format = '0.00'
                    cell.font = data_font
                    cell.alignment = right_align
                    continue
                except (TypeError, ValueError):
                    pass

            if key in upside_keys and value is not None:
                try:
                    cell.value = round(float(value), 1)
                    cell.number_format = '#,##0.0'
                    cell.font = data_font
                    cell.alignment = right_align
                    continue
                except (TypeError, ValueError):
                    pass

            if key == "3M MV Change ($MM)" and value is not None:
                try:
                    v = float(value)
                    cell.value = round(v, 0)
                    cell.number_format = '+#,##0;-#,##0;"-"'
                    cell.font = Font(name="Calibri", size=9, bold=True,
                                     color=_GREEN if v > 0 else (_RED if v < 0 else "5D6A7D"))
                    cell.alignment = right_align
                    continue
                except (TypeError, ValueError):
                    pass

            if key == "Issuer" and is_holding:
                name = str(value) if value is not None else ""
                cell.value = CellRichText(
                    TextBlock(InlineFont(b=True, sz=18, color="14233B"), name),
                    TextBlock(InlineFont(b=True, sz=13, vertAlign="superscript", color=_GREEN), "H"),
                )
                cell.alignment = left_align
                continue
            cell.value = value if value is not None else ""
            cell.font = data_font_bold if key == "Issuer" else data_font
            cell.alignment = left_align if key in ("Issuer", "Sector") else right_align

    # ── Column widths ─────────────────────────────────────────────────────
    col_widths = [28, 14, 9, 10, 9, 9, 10, 9, 7, 7, 8, 8, 10]
    for c, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # freeze panes below header rows
    ws.freeze_panes = "A3"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="issuer_screen.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/", response_class=FileResponse)
def root() -> FileResponse:
    return FileResponse(Path(__file__).parent / "static" / "index.html")


@app.post("/login")
def login(username: str = Form(...), password: str = Form(...)) -> JSONResponse:
    if username != settings.username or password != settings.password:
        raise HTTPException(status_code=401, detail="Invalid credentials")
    token = create_session_token(settings.app_secret, username)
    response = JSONResponse({"ok": True, "username": username})
    response.set_cookie(
        key=SESSION_COOKIE,
        value=token,
        httponly=True,
        samesite="lax",
        secure=settings.cookie_secure,
        max_age=60 * 60 * 12,
    )
    return response


@app.post("/logout")
def logout() -> JSONResponse:
    response = JSONResponse({"ok": True})
    response.delete_cookie(SESSION_COOKIE)
    return response


@app.get("/api/session")
def session(request: Request) -> JSONResponse:
    token = request.cookies.get(SESSION_COOKIE)
    user = decode_session_token(settings.app_secret, token)
    if not user:
        return JSONResponse({"authenticated": False})
    return JSONResponse({"authenticated": True, "username": user.username})


_market_cache: dict = {"mtime": None, "data": None}


def ensure_market_loaded() -> dict:
    path = settings.market_data_path
    if not path.exists():
        raise HTTPException(status_code=500, detail=f"Market data not found: {path}")
    mtime = path.stat().st_mtime
    if _market_cache["data"] is None or _market_cache["mtime"] != mtime:
        _market_cache["data"] = load_market_data(path)
        _market_cache["mtime"] = mtime
    return _market_cache["data"]


@app.get("/api/market")
def market(_: str = Depends(get_current_user)) -> JSONResponse:
    return JSONResponse(ensure_market_loaded())


@app.get("/api/dashboard")
def dashboard(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    return JSONResponse(enrich_issuers_with_reports(dataset))


@app.get("/api/issuer-detail")
def issuer_detail(parent_ticker: str, _: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    rows = [row for row in dataset.instrument_rows if row.get("PARENT_TICKER") == parent_ticker]
    if not rows:
        raise HTTPException(status_code=404, detail="Issuer not found")
    issuer_name = next((row.get("Issuer") for row in dataset.issuer_table if row.get("PARENT_TICKER") == parent_ticker), "")
    report = report_service.snapshot().get(normalize_company_name(str(issuer_name)))
    return JSONResponse({"issuer": parent_ticker, "rows": rows, "report": report})


@app.get("/api/price-movers")
def price_movers(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    issuer_name_map = {
        row.get("PARENT_TICKER"): row.get("Issuer")
        for row in dataset.issuer_table
    }

    movers: list[dict] = []
    for row in dataset.instrument_rows:
        current_px = pd.to_numeric(row.get("PX_MID"), errors="coerce")
        if pd.isna(current_px):
            continue
        move_3m = row.get("PRICE_MOVE_3M")
        move_3m_val = float(move_3m) if move_3m is not None and not pd.isna(pd.to_numeric(move_3m, errors="coerce")) else None
        move_7d = row.get("PRICE_MOVE_7D")
        move_7d_val = float(move_7d) if move_7d is not None and not pd.isna(pd.to_numeric(move_7d, errors="coerce")) else None
        qualifies_3m = move_3m_val is not None and abs(move_3m_val) > 10
        qualifies_7d = move_7d_val is not None and abs(move_7d_val) > 3
        if not qualifies_3m and not qualifies_7d:
            continue
        raw_vol = row.get("LAST_30D_VOLUME_MM")
        mover = {
            "Issuer Name": issuer_name_map.get(row.get("PARENT_TICKER")) or row.get("PARENT_TICKER"),
            "Security Name": row.get("NAME"),
            "Sector": row.get("SECTOR"),
            "Lien": row.get("PAYMENT_RANK"),
            "Amount Out ($MM)": row.get("AMT_OUTSTANDING_MM"),
            "Last Month Traded Volume ($MM)": round(float(raw_vol), 4) if raw_vol is not None and not pd.isna(raw_vol) else None,
            "Current Px": current_px,
            "3M Price Move": round(move_3m_val, 4) if move_3m_val is not None else None,
            "7D Price Move": round(move_7d_val, 4) if move_7d_val is not None else None,
            "PX_LOW_52W": row.get("PX_LOW_52W"),
            "PX_HIGH_52W": row.get("PX_HIGH_52W"),
        }
        movers.append(mover)

    sector_df = pd.DataFrame(movers)
    if not sector_df.empty:
        sector_summary = (
            sector_df.groupby("Sector", dropna=False)["Amount Out ($MM)"]
            .sum(min_count=1)
            .reset_index()
            .fillna({"Sector": "Unknown"})
            .sort_values("Amount Out ($MM)", ascending=False, na_position="last")
        )
        sector_bars = sector_summary.to_dict(orient="records")
    else:
        sector_bars = []

    return JSONResponse({"rows": movers, "sector_summary": sector_bars})


@app.get("/api/abnormal-prices")
def abnormal_prices(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    return JSONResponse({"rows": dataset.abnormal_price_rows})


@app.get("/api/excluded")
def excluded_securities(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    return JSONResponse({"rows": dataset.excluded_rows})


@app.get("/api/loans")
def loans(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    return JSONResponse({"rows": dataset.loan_rows})


@app.get("/api/bonds")
def bonds(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    return JSONResponse({"rows": dataset.bond_rows})


@app.get("/api/coverage")
def get_coverage(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    return JSONResponse({"coverages": _build_coverage_map(dataset)})


@app.post("/api/coverage")
def post_coverage(payload: dict = Body(...), _: str = Depends(get_current_user)) -> JSONResponse:
    ticker = payload.get("ticker", "")
    primary = payload.get("primary", [])
    secondary = payload.get("secondary", [])
    if not ticker:
        raise HTTPException(status_code=400, detail="ticker required")
    save_coverage(str(ticker), list(primary), list(secondary))
    return JSONResponse({"ok": True})


@app.post("/api/export/issuers")
def export_issuers(payload: dict = Body(...), _: str = Depends(get_current_user)) -> StreamingResponse:
    rows = payload.get("rows", [])
    upside_mode = str(payload.get("upsideMode", "52w"))
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows must be a list")
    return build_issuer_excel_response(rows, upside_mode)


def build_bonds_excel_response(rows: list[dict], metadata: dict | None = None) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Bond Instruments"

    anchor_date = (metadata or {}).get("anchor_date", "")

    # (data_key, header_label, col_width, h_align)
    COLS = [
        ("#",                "#",                4,  "center"),
        ("PARENT_TICKER",    "Ticker",           7,  "center"),
        ("Issuer",           "Issuer Name",     22,  "left"),
        ("NAME",             "Bond Description",30,  "left"),
        ("COUPON_RATE",      "Coupon",           7,  "center"),
        ("PAYMENT_RANK",     "Lien",             7,  "center"),
        ("SECTOR",           "Sector",          11,  "left"),
        ("AMT_OUTSTANDING_MM","Size ($MM)",       9,  "right"),
        ("MATURITY",         "Maturity",         9,  "center"),
        ("PX_MID",           "Mid Price",        8,  "right"),
        ("YIELD",            "Yld to Mty",       8,  "right"),
        ("PRICE_MOVE_3M",    "3M Px Chg",        8,  "right"),
        ("PRICE_MOVE_7D",    "7D Px Chg",        8,  "right"),
        ("MV_CHANGE_3M_MM",  "3M MV ($MM)",      9,  "right"),
        ("MV_CHANGE_7D_MM",  "7D MV ($MM)",      9,  "right"),
        ("PX_LOW_52W",       "52W Low",          7,  "right"),
        ("PX_HIGH_52W",      "52W High",         7,  "right"),
    ]
    n_cols = len(COLS)

    hdr_fill = PatternFill("solid", fgColor=_LOAN_NAVY)
    alt_fill = PatternFill("solid", fgColor=_LOAN_ALT_ROW)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    data_font = Font(name="Calibri", size=9)
    bold_font = Font(name="Calibri", size=9, bold=True)

    c_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_al = Alignment(horizontal="left", vertical="center")
    r_al = Alignment(horizontal="right", vertical="center")
    def _al(h): return c_al if h == "center" else (l_al if h == "left" else r_al)

    no_s = Side(style=None)
    thin_bot = Side(style="thin", color="B8CCE4")
    div_l = Side(style="thin", color=_LOAN_DIVIDER)

    GROUPS = [
        (1,  7,  ""),
        (8,  11, "Illustrative"),
        (12, 13, "Price Movement"),
        (14, 15, "MV Change ($MM)"),
        (16, 17, "52W Range"),
    ]
    group_starts = {g[0] for g in GROUPS if g[2]}

    title_text = f"Bond Instruments{'  |  As of ' + anchor_date if anchor_date else ''}"
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    tc.fill = hdr_fill
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    for col in range(2, n_cols + 1):
        ws.cell(row=1, column=col).fill = hdr_fill
    ws.row_dimensions[1].height = 22

    for start, end, label in GROUPS:
        gc = ws.cell(row=2, column=start, value=label)
        gc.font = hdr_font
        gc.fill = hdr_fill
        gc.alignment = c_al
        if end > start:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
    for col in range(1, n_cols + 1):
        gc = ws.cell(row=2, column=col)
        if gc.value is None:
            gc.fill = hdr_fill
        gc.border = Border(
            bottom=Side(style="thin", color="4A6FA5"),
            left=div_l if col in group_starts else no_s,
            right=no_s, top=no_s,
        )
    ws.row_dimensions[2].height = 14

    for c_idx, (key, label, width, h) in enumerate(COLS, start=1):
        lc = ws.cell(row=3, column=c_idx, value=label)
        lc.font = hdr_font
        lc.fill = hdr_fill
        lc.alignment = c_al
        lc.border = Border(
            bottom=Side(style="medium", color=_LOAN_DIVIDER),
            left=div_l if c_idx in group_starts else no_s,
            right=no_s, top=no_s,
        )
    ws.row_dimensions[3].height = 20

    move_keys = {"PRICE_MOVE_3M", "PRICE_MOVE_7D"}
    mv_keys = {"MV_CHANGE_3M_MM", "MV_CHANGE_7D_MM"}

    for r_idx, row in enumerate(rows, start=1):
        xl_row = r_idx + 3
        is_holding = bool(row.get("_IS_HOLDING"))
        if is_holding:
            row_fill = PatternFill("solid", fgColor="C8E8D5") if r_idx % 2 == 0 else PatternFill("solid", fgColor="DFF2E9")
        else:
            row_fill = alt_fill if r_idx % 2 == 0 else white_fill
        for c_idx, (key, label, width, h) in enumerate(COLS, start=1):
            dc = ws.cell(row=xl_row, column=c_idx)
            dc.fill = row_fill
            dc.border = Border(
                bottom=thin_bot,
                left=div_l if c_idx in group_starts else no_s,
                right=no_s, top=no_s,
            )

            if key == "#":
                dc.value = r_idx
                dc.font = data_font
                dc.alignment = c_al
                continue

            value = row.get(key)

            if key == "COUPON_RATE" and value is not None:
                try:
                    dc.value = round(float(value), 2)
                    dc.number_format = '0.00"%"'
                    dc.font = data_font
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key in move_keys and value is not None:
                try:
                    v = float(value)
                    dc.value = round(v, 2)
                    dc.number_format = "+0.00;-0.00;-"
                    dc.font = Font(name="Calibri", size=9, bold=True,
                                   color=_GREEN if v > 0 else (_RED if v < 0 else "808080"))
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key in mv_keys and value is not None:
                try:
                    v = float(value)
                    dc.value = round(v, 0)
                    dc.number_format = '+#,##0;-#,##0;"-"'
                    dc.font = Font(name="Calibri", size=9, bold=True,
                                   color=_GREEN if v > 0 else (_RED if v < 0 else "808080"))
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key == "YIELD" and value is not None:
                try:
                    dc.value = round(float(value), 1)
                    dc.number_format = '0.0"%"'
                    dc.font = data_font
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key in ("PX_MID", "PX_LOW_52W", "PX_HIGH_52W") and value is not None:
                try:
                    dc.value = round(float(value), 3)
                    dc.number_format = "0.000"
                    dc.font = data_font
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key == "AMT_OUTSTANDING_MM" and value is not None:
                try:
                    dc.value = round(float(value), 0)
                    dc.number_format = '"$"#,##0'
                    dc.font = data_font
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            dc.value = value if value is not None else ""
            dc.font = bold_font if key == "PARENT_TICKER" else data_font
            dc.alignment = _al(h)

    for c_idx, (key, label, width, h) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.freeze_panes = "D4"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    hdrs = {"Content-Disposition": 'attachment; filename="bonds.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=hdrs,
    )


@app.post("/api/export/bonds")
def export_bonds(payload: dict = Body(...), _: str = Depends(get_current_user)) -> StreamingResponse:
    rows = payload.get("rows", [])
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows must be a list")
    dataset = ensure_data_loaded()
    return build_bonds_excel_response(rows, dataset.metadata)


@app.post("/api/export/detail")
def export_detail(payload: dict = Body(...), _: str = Depends(get_current_user)) -> StreamingResponse:
    rows = payload.get("rows", [])
    issuer = str(payload.get("issuer") or "detail_view").strip() or "detail_view"
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows must be a list")
    safe_name = "".join(ch if ch.isalnum() or ch in ("-", "_") else "_" for ch in issuer)[:80]
    return build_excel_response(rows, f"{safe_name}.xlsx", "Detail")


# ---------------------------------------------------------------------------
# Loans export – themed (dark navy / light blue, matching PDF package aesthetic)
# ---------------------------------------------------------------------------
_LOAN_NAVY = "1F3864"
_LOAN_ALT_ROW = "D9E1F2"
_LOAN_DIVIDER = "4472C4"

def build_loans_excel_response(rows: list[dict], metadata: dict | None = None) -> StreamingResponse:
    wb = Workbook()
    ws = wb.active
    ws.title = "Loan Instruments"

    anchor_date = (metadata or {}).get("anchor_date", "")

    # (data_key, header_label, col_width, h_align)
    COLS = [
        ("#",                "#",                4,  "center"),
        ("PARENT_TICKER",    "Ticker",           7,  "center"),
        ("Issuer",           "Issuer Name",     22,  "left"),
        ("NAME",             "Loan Description",30,  "left"),
        ("LOAN_TYPE",        "Type",             7,  "center"),
        ("PAYMENT_RANK",     "Lien",             7,  "center"),
        ("SECTOR",           "Sector",          11,  "left"),
        ("AMT_OUTSTANDING_MM","Size ($MM)",       9,  "right"),
        ("MATURITY",         "Maturity",         9,  "center"),
        ("PX_MID",           "Mid Price",        8,  "right"),
        ("YIELD",            "Yld to Mty",       8,  "right"),
        ("PRICE_MOVE_3M",    "3M Px Chg",        8,  "right"),
        ("PRICE_MOVE_7D",    "7D Px Chg",        8,  "right"),
        ("MV_CHANGE_3M_MM",  "3M MV ($MM)",      9,  "right"),
        ("MV_CHANGE_7D_MM",  "7D MV ($MM)",      9,  "right"),
        ("PX_LOW_52W",       "52W Low",          7,  "right"),
        ("PX_HIGH_52W",      "52W High",         7,  "right"),
    ]
    n_cols = len(COLS)

    hdr_fill = PatternFill("solid", fgColor=_LOAN_NAVY)
    alt_fill = PatternFill("solid", fgColor=_LOAN_ALT_ROW)
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=9)
    data_font = Font(name="Calibri", size=9)
    bold_font = Font(name="Calibri", size=9, bold=True)

    c_al = Alignment(horizontal="center", vertical="center", wrap_text=True)
    l_al = Alignment(horizontal="left", vertical="center")
    r_al = Alignment(horizontal="right", vertical="center")
    def _al(h): return c_al if h == "center" else (l_al if h == "left" else r_al)

    no_s = Side(style=None)
    thin_bot = Side(style="thin", color="B8CCE4")
    div_l = Side(style="thin", color=_LOAN_DIVIDER)

    # Groups for row 2: (start_col, end_col, label)
    GROUPS = [
        (1,  7,  ""),
        (8,  11, "Illustrative"),
        (12, 13, "Price Movement"),
        (14, 15, "MV Change ($MM)"),
        (16, 17, "52W Range"),
    ]
    group_starts = {g[0] for g in GROUPS if g[2]}

    # ── Row 1: title banner ───────────────────────────────────────────────
    title_text = f"Loan Instruments{'  |  As of ' + anchor_date if anchor_date else ''}"
    tc = ws.cell(row=1, column=1, value=title_text)
    tc.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    tc.fill = hdr_fill
    tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    for col in range(2, n_cols + 1):
        ws.cell(row=1, column=col).fill = hdr_fill
    ws.row_dimensions[1].height = 22

    # ── Row 2: group headers ──────────────────────────────────────────────
    for start, end, label in GROUPS:
        gc = ws.cell(row=2, column=start, value=label)
        gc.font = hdr_font
        gc.fill = hdr_fill
        gc.alignment = c_al
        if end > start:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
    for col in range(1, n_cols + 1):
        gc = ws.cell(row=2, column=col)
        if gc.value is None:
            gc.fill = hdr_fill
        gc.border = Border(
            bottom=Side(style="thin", color="4A6FA5"),
            left=div_l if col in group_starts else no_s,
            right=no_s, top=no_s,
        )
    ws.row_dimensions[2].height = 14

    # ── Row 3: column labels ──────────────────────────────────────────────
    for c_idx, (key, label, width, h) in enumerate(COLS, start=1):
        lc = ws.cell(row=3, column=c_idx, value=label)
        lc.font = hdr_font
        lc.fill = hdr_fill
        lc.alignment = c_al
        lc.border = Border(
            bottom=Side(style="medium", color=_LOAN_DIVIDER),
            left=div_l if c_idx in group_starts else no_s,
            right=no_s, top=no_s,
        )
    ws.row_dimensions[3].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────
    move_keys = {"PRICE_MOVE_3M", "PRICE_MOVE_7D"}
    mv_keys = {"MV_CHANGE_3M_MM", "MV_CHANGE_7D_MM"}

    for r_idx, row in enumerate(rows, start=1):
        xl_row = r_idx + 3
        row_fill = alt_fill if r_idx % 2 == 0 else white_fill
        for c_idx, (key, label, width, h) in enumerate(COLS, start=1):
            dc = ws.cell(row=xl_row, column=c_idx)
            dc.fill = row_fill
            dc.border = Border(
                bottom=thin_bot,
                left=div_l if c_idx in group_starts else no_s,
                right=no_s, top=no_s,
            )

            if key == "#":
                dc.value = r_idx
                dc.font = data_font
                dc.alignment = c_al
                continue

            value = row.get(key)

            if key in move_keys and value is not None:
                try:
                    v = float(value)
                    dc.value = round(v, 2)
                    dc.number_format = "+0.00;-0.00;-"
                    dc.font = Font(name="Calibri", size=9, bold=True,
                                   color=_GREEN if v > 0 else (_RED if v < 0 else "808080"))
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key in mv_keys and value is not None:
                try:
                    v = float(value)
                    dc.value = round(v, 0)
                    dc.number_format = '+#,##0;-#,##0;"-"'
                    dc.font = Font(name="Calibri", size=9, bold=True,
                                   color=_GREEN if v > 0 else (_RED if v < 0 else "808080"))
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key == "YIELD" and value is not None:
                try:
                    dc.value = round(float(value), 1)
                    dc.number_format = '0.0"%"'
                    dc.font = data_font
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key in ("PX_MID", "PX_LOW_52W", "PX_HIGH_52W") and value is not None:
                try:
                    dc.value = round(float(value), 3)
                    dc.number_format = "0.000"
                    dc.font = data_font
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            if key == "AMT_OUTSTANDING_MM" and value is not None:
                try:
                    dc.value = round(float(value), 0)
                    dc.number_format = '"$"#,##0'
                    dc.font = data_font
                    dc.alignment = r_al
                    continue
                except (TypeError, ValueError):
                    pass

            dc.value = value if value is not None else ""
            dc.font = bold_font if key == "PARENT_TICKER" else data_font
            dc.alignment = _al(h)

    # ── Column widths + freeze ────────────────────────────────────────────
    for c_idx, (key, label, width, h) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(c_idx)].width = width
    ws.freeze_panes = "D4"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    hdrs = {"Content-Disposition": 'attachment; filename="loans.xlsx"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=hdrs,
    )


@app.post("/api/export/loans")
def export_loans(payload: dict = Body(...), _: str = Depends(get_current_user)) -> StreamingResponse:
    rows = payload.get("rows", [])
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows must be a list")
    dataset = ensure_data_loaded()
    return build_loans_excel_response(rows, dataset.metadata)


@app.post("/api/admin/reload")
def admin_reload(_: str = Depends(get_current_user)) -> JSONResponse:
    workbook_path = settings.workbook_path
    if not workbook_path.exists():
        raise HTTPException(status_code=500, detail=f"Workbook not found: {workbook_path}")
    dataset = store.refresh(workbook_path)
    return JSONResponse({"ok": True, "metadata": dataset.metadata})


@app.get("/api/debug/instrument-columns")
def debug_instrument_columns(_: str = Depends(get_current_user)) -> JSONResponse:
    dataset = ensure_data_loaded()
    if not dataset.instrument_rows:
        return JSONResponse({"columns": []})
    all_keys = sorted(dataset.instrument_rows[0].keys())
    lqa_keys = [k for k in all_keys if "LQA" in k.upper() or "VOLUME" in k.upper() or "DAILY" in k.upper()]
    return JSONResponse({"lqa_and_volume_keys": lqa_keys, "all_keys": all_keys})


@app.post("/api/admin/upload")
def admin_upload(workbook: UploadFile = File(...), _: str = Depends(get_current_user)) -> JSONResponse:
    if not workbook.filename or not workbook.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx workbook")

    target_path = DATA_DIR / "Master_File.xlsx"
    with target_path.open("wb") as target:
        shutil.copyfileobj(workbook.file, target)

    dataset = store.refresh(target_path)
    return JSONResponse({"ok": True, "metadata": dataset.metadata})


@app.post("/api/admin/process-reports")
def admin_process_reports(_: str = Depends(get_current_user)) -> JSONResponse:
    try:
        records = report_service.refresh_once()
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Report processing failed: {exc}") from exc
    return JSONResponse({"ok": True, "processed_report_count": len(records)})


app.mount("/static", StaticFiles(directory=Path(__file__).parent / "static"), name="static")
